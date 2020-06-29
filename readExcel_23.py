'''
Agenda
   Excel operations:
   1.)Reading data from Excel File.
   2.)Writing data in Excel file
   3.)Data Driven Test Case

'''
import openpyxl
path="C:\\Users\\Dell\\PycharmProjects\\selenium\\data\\emp.xlsx"

workbook=openpyxl.load_workbook(path)
sheet=workbook.active  #workbook.get_sheet_by_name("Sheet1")
rows=sheet.max_row
cols=sheet.max_column
print(rows,cols) #13,4
for r in range(1,rows+1):
    # print(r)
   for c in range(1,cols+1):
       # print(c)
       print(sheet.cell(row=r,column=c).value,end="            ")
   print()










